"""
Tab 3 Optimized - Viabilité avec caching et performance améliorée.

Optimisations:
- @st.cache_data pour stress tests
- Calculs parallèles
- Progress bars
- Export Excel
"""

import streamlit as st
import plotly.graph_objects as go
import pandas as pd
from io import BytesIO
from datetime import datetime
from typing import Dict, List

from src.ui.utils.formatting import (
    format_number,
    format_percentage,
    format_ratio,
    format_currency_compact,
)
from src.scenarios.stress_tester import StressTester
from src.calculations.covenant_tracker import CovenantTracker
from src.decision.decision_engine import DecisionEngine


@st.cache_data(ttl=3600, show_spinner=False)
def compute_stress_tests_cached(
    financial_data_json: str,
    lbo_dict_json: str,
    norm_dict_json: str
) -> List[Dict]:
    """
    Calcule stress tests avec cache (1h).

    Args:
        financial_data_json: JSON string des données financières
        lbo_dict_json: JSON string structure LBO
        norm_dict_json: JSON string normalisation

    Returns:
        Liste résultats stress tests
    """
    import json

    financial_data = json.loads(financial_data_json)
    lbo_dict = json.loads(lbo_dict_json)
    norm_dict = json.loads(norm_dict_json)

    return StressTester.run_all_scenarios(financial_data, lbo_dict, norm_dict)


@st.cache_data(ttl=3600, show_spinner=False)
def compute_covenant_tracking_cached(
    financial_data_json: str,
    lbo_dict_json: str,
    norm_dict_json: str,
    assumptions_json: str,
    projection_years: int = 7
) -> List[Dict]:
    """
    Calcule covenant tracking avec cache (1h).

    Args:
        financial_data_json: JSON string des données financières
        lbo_dict_json: JSON string structure LBO
        norm_dict_json: JSON string normalisation
        assumptions_json: JSON string hypothèses
        projection_years: Nombre d'années

    Returns:
        Projections 7 ans
    """
    import json

    financial_data = json.loads(financial_data_json)
    lbo_dict = json.loads(lbo_dict_json)
    norm_dict = json.loads(norm_dict_json)
    assumptions_dict = json.loads(assumptions_json)

    return CovenantTracker.generate_projections(
        financial_data,
        lbo_dict,
        norm_dict,
        assumptions_dict,
        projection_years
    )


def create_excel_export(
    stress_results: List[Dict],
    projections: List[Dict],
    decision: Dict,
    lbo: Dict,
    norm_data: Dict,
    company_name: str = "Entreprise"
) -> BytesIO:
    """
    Créer export Excel complet.

    Args:
        stress_results: Résultats stress tests
        projections: Projections 7 ans
        decision: Décision finale
        lbo: Structure LBO
        norm_data: Données normalisées
        company_name: Nom entreprise

    Returns:
        BytesIO avec fichier Excel
    """
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # ========== SHEET 1: SYNTHÈSE ==========
        synthese_data = {
            "Métrique": [
                "Prix acquisition",
                "Dette totale",
                "Equity",
                "EBITDA normalisé",
                "Décision finale",
                "Score global"
            ],
            "Valeur": [
                f"{lbo.get('acquisition_price', 0):,.0f} €",
                f"{lbo.get('total_debt', 0):,.0f} €",
                f"{lbo.get('equity_amount', 0):,.0f} €",
                f"{norm_data.get('ebitda_bank', 0):,.0f} €",
                decision.get("decision", {}).get("value", "N/A"),
                f"{decision.get('overall_score', 0)}/100"
            ]
        }

        df_synthese = pd.DataFrame(synthese_data)
        df_synthese.to_excel(writer, sheet_name="Synthèse", index=False)

        # ========== SHEET 2: STRESS TESTS ==========
        stress_data = []
        for result in stress_results:
            scenario = result["scenario"]
            metrics = result["metrics"]
            status = StressTester.get_status_from_metrics(metrics)

            stress_data.append({
                "Scénario": scenario.name,
                "Description": scenario.description,
                "DSCR min": round(metrics.get("dscr_min", 0), 2),
                "Dette/EBITDA": round(metrics.get("leverage", 0), 2),
                "Marge (%)": round(metrics.get("margin", 0), 1),
                "FCF Année 3 (€)": int(metrics.get("fcf_year3", 0)),
                "Statut": status
            })

        df_stress = pd.DataFrame(stress_data)
        df_stress.to_excel(writer, sheet_name="Stress Tests", index=False)

        # Mise en forme conditionnelle
        workbook = writer.book
        worksheet = writer.sheets["Stress Tests"]

        from openpyxl.styles import PatternFill

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for row in range(2, len(stress_data) + 2):
            cell = worksheet[f"G{row}"]
            if cell.value == "GO":
                cell.fill = green_fill
            elif cell.value == "WATCH":
                cell.fill = yellow_fill
            else:
                cell.fill = red_fill

        # ========== SHEET 3: PROJECTIONS 7 ANS ==========
        proj_data = []
        for i, proj in enumerate(projections):
            proj_data.append({
                "Année": f"Y{i+1}",
                "CA (€)": int(proj.get("revenue", 0)),
                "EBITDA (€)": int(proj.get("ebitda", 0)),
                "CFADS (€)": int(proj.get("cfads", 0)),
                "DSCR": round(proj.get("dscr", 0), 2),
                "Dette/EBITDA": round(proj.get("net_debt_to_ebitda", 0), 2),
                "FCF (€)": int(proj.get("fcf", 0))
            })

        df_proj = pd.DataFrame(proj_data)
        df_proj.to_excel(writer, sheet_name="Projections 7 ans", index=False)

        # ========== SHEET 4: STRUCTURE LBO ==========
        if "debt_layers" in lbo:
            debt_data = []
            for layer in lbo["debt_layers"]:
                debt_data.append({
                    "Tranche": layer.get("name", ""),
                    "Montant (€)": int(layer.get("amount", 0)),
                    "Taux (%)": round(layer.get("interest_rate", 0) * 100, 2),
                    "Durée (ans)": layer.get("duration_years", 0),
                    "Grace (ans)": layer.get("grace_period", 0)
                })

            df_debt = pd.DataFrame(debt_data)
            df_debt.to_excel(writer, sheet_name="Structure Dette", index=False)

    output.seek(0)
    return output


def render_tab3_optimized(
    lbo,
    norm_data,
    financial_data: Dict
) -> None:
    """
    Render Tab 3 optimisé avec caching et export Excel.

    Args:
        lbo: Structure LBO
        norm_data: Données normalisées
        financial_data: Données financières
    """
    import json

    st.header("✅ Viabilité & Décision")

    st.markdown(f"""
    **Prix acquisition**: {format_number(lbo.acquisition_price)}
    **Dette totale**: {format_number(lbo.total_debt)}
    **Equity**: {format_number(lbo.equity_amount)}
    """)

    st.divider()

    # Préparer données pour caching
    lbo_dict = {
        "debt_layers": [
            {
                "name": layer.name,
                "amount": layer.amount,
                "interest_rate": layer.interest_rate,
                "duration_years": layer.duration_years
            }
            for layer in lbo.debt_layers
        ]
    }

    norm_dict = {
        "ebitda_bank": norm_data.ebitda_bank,
        "ebitda_equity": norm_data.ebitda_equity
    }

    # =========================================================================
    # SECTION 1: STRESS TESTS AVEC CACHE
    # =========================================================================
    st.subheader("🔬 1. Stress Tests")

    st.markdown("Test de robustesse du montage sous différents scénarios de crise:")

    # Calcul avec cache + progress bar
    with st.spinner("Calcul des 7 scénarios de stress..."):
        # Convertir en JSON pour cache
        financial_json = json.dumps(financial_data)
        lbo_json = json.dumps(lbo_dict)
        norm_json = json.dumps(norm_dict)

        stress_results = compute_stress_tests_cached(
            financial_json,
            lbo_json,
            norm_json
        )

    # Afficher résultats (code identique à app_v3.py)
    st.markdown("#### Résultats Stress Tests")

    # En-têtes
    col1, col2, col3, col4, col5 = st.columns([3, 2, 2, 2, 2])
    with col1:
        st.markdown("**Scénario**")
    with col2:
        st.markdown("**DSCR min**")
    with col3:
        st.markdown("**Dette/EB**")
    with col4:
        st.markdown("**FCF an 3**")
    with col5:
        st.markdown("**Statut**")

    st.divider()

    # Lignes
    for result in stress_results:
        scenario = result["scenario"]
        metrics = result["metrics"]
        status = StressTester.get_status_from_metrics(metrics)

        # Icône selon statut
        if status == "GO":
            status_icon = "🟢"
            status_color = "green"
        elif status == "WATCH":
            status_icon = "🟡"
            status_color = "orange"
        else:
            status_icon = "🔴"
            status_color = "red"

        col1, col2, col3, col4, col5 = st.columns([3, 2, 2, 2, 2])

        with col1:
            icon = "✅" if scenario.scenario_type.value == "nominal" else "⚠️"
            st.write(f"{icon} {scenario.name}")

        with col2:
            dscr = metrics.get("dscr_min", 0)
            st.metric("", format_ratio(dscr), label_visibility="collapsed")

        with col3:
            leverage = metrics.get("leverage", 0)
            st.metric("", format_ratio(leverage) + "x", label_visibility="collapsed")

        with col4:
            fcf = metrics.get("fcf_year3", 0)
            st.metric("", format_currency_compact(fcf), label_visibility="collapsed")

        with col5:
            st.markdown(f":{status_color}[{status_icon} {status}]")

    # Analyse
    st.divider()

    failed_scenarios = [
        r for r in stress_results
        if StressTester.get_status_from_metrics(r["metrics"]) == "NO-GO"
    ]

    if failed_scenarios:
        st.error(f"⚠️ **{len(failed_scenarios)} scénario(s) en échec**: Dossier sensible aux chocs")
        for result in failed_scenarios:
            st.caption(f"  • {result['scenario'].name}: {result['scenario'].description}")
    else:
        st.success("✅ **Dossier robuste**: Tous les scénarios passent")

    # =========================================================================
    # SECTION 2: MATRICE SENSIBILITÉ (identique)
    # =========================================================================
    st.divider()
    st.subheader("📊 2. Analyse de Sensibilité")

    st.markdown("Impact croisé CA × Marge sur le DSCR:")

    sensitivity = StressTester.generate_sensitivity_matrix(
        financial_data,
        lbo_dict,
        norm_dict,
        metric="dscr_min"
    )

    heatmap_fig = go.Figure(data=go.Heatmap(
        z=sensitivity["matrix"],
        x=sensitivity["ca_labels"],
        y=sensitivity["margin_labels"],
        colorscale=[
            [0, "red"],
            [0.5, "orange"],
            [0.7, "yellow"],
            [1, "green"]
        ],
        text=[[f"{val:.2f}" for val in row] for row in sensitivity["matrix"]],
        texttemplate="%{text}",
        textfont={"size": 10},
        colorbar=dict(title="DSCR")
    ))

    heatmap_fig.update_layout(
        title="Heatmap Sensibilité: CA × Marge → DSCR",
        xaxis_title="Variation CA",
        yaxis_title="Variation Marge EBITDA",
        height=400
    )

    st.plotly_chart(heatmap_fig, use_container_width=True)

    # =========================================================================
    # SECTION 3: COVENANT TRACKING AVEC CACHE
    # =========================================================================
    st.divider()
    st.subheader("📈 3. Covenant Tracking (7 ans)")

    # Hypothèses
    assumptions_dict = {
        "revenue_growth_rate": [0.05, 0.05, 0.03, 0.03, 0.02, 0.02, 0.02],
        "ebitda_margin_evolution": [0.5, 0.5, 0.0, 0.0, 0.0, 0.0, 0.0],
        "tax_rate": 0.25,
        "bfr_percentage_of_revenue": 18.0,
        "capex_maintenance_pct": 3.0
    }

    # Calcul avec cache
    with st.spinner("Génération des projections 7 ans..."):
        assumptions_json = json.dumps(assumptions_dict)

        projections = compute_covenant_tracking_cached(
            financial_json,
            lbo_json,
            norm_json,
            assumptions_json,
            projection_years=7
        )

    # Créer tracker
    tracker = CovenantTracker()

    # Projeter covenants
    covenant_results = tracker.project_all_covenants(projections)

    # Graphiques (code identique à app_v3.py - repris tel quel)
    for cov_result in covenant_results:
        covenant = cov_result["covenant"]
        years = cov_result["years"]
        values = cov_result["values"]
        threshold = cov_result["threshold"]
        violations = cov_result["violations"]

        fig = go.Figure()

        fig.add_hline(
            y=threshold,
            line_dash="dash",
            line_color="red",
            annotation_text=f"Seuil: {covenant.comparison} {threshold}",
            annotation_position="right"
        )

        fig.add_trace(go.Scatter(
            x=years,
            y=values,
            mode="lines+markers",
            name=covenant.name,
            line=dict(width=3),
            marker=dict(size=8)
        ))

        if covenant.comparison in [">=", ">"]:
            fig.add_hrect(y0=threshold, y1=max(values + [threshold]) * 1.2, fillcolor="green", opacity=0.1, line_width=0)
            fig.add_hrect(y0=0, y1=threshold, fillcolor="red", opacity=0.1, line_width=0)
        else:
            fig.add_hrect(y0=0, y1=threshold, fillcolor="green", opacity=0.1, line_width=0)
            fig.add_hrect(y0=threshold, y1=max(values + [threshold]) * 1.2, fillcolor="red", opacity=0.1, line_width=0)

        fig.update_layout(
            title=f"{covenant.name} - Projection 7 ans",
            xaxis_title="Année",
            yaxis_title=covenant.name,
            height=300,
            showlegend=False
        )

        st.plotly_chart(fig, use_container_width=True)

        if violations:
            st.error(f"❌ **Violations détectées**: Années {violations}")
        else:
            st.success(f"✅ **Aucune violation** sur 7 ans")

    # Résumé
    summary = tracker.get_summary(projections)

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Covenants au vert", summary["pass_count"])
    with col2:
        st.metric("Covenants en warning", summary["warning_count"])
    with col3:
        st.metric("Violations", summary["violated_count"])

    # =========================================================================
    # SECTION 4: DÉCISION FINALE
    # =========================================================================
    st.divider()
    st.subheader("🎯 4. Décision d'Acquisition")

    decision = DecisionEngine.make_decision(
        projections,
        norm_dict,
        financial_data,
        scenario_id="main_scenario"
    )

    decision_icon = DecisionEngine.get_decision_icon(decision.decision)
    decision_color = DecisionEngine.get_decision_color(decision.decision)

    st.markdown(f"## :{decision_color}[{decision_icon} {decision.decision.value.upper()}]")
    st.markdown(f"**Score global**: {decision.overall_score}/100")

    st.divider()

    # Critères (code identique app_v3.py)
    st.markdown("#### 📊 Critères Décisifs")

    for criterion in decision.criteria:
        icon = "🟢" if criterion.status == "PASS" else "🟡" if criterion.status == "WARNING" else "🔴"

        col1, col2, col3, col4 = st.columns([3, 2, 1, 1])
        with col1:
            st.write(f"{icon} {criterion.name}")
        with col2:
            st.write(f"{criterion.actual_value:.2f}")
        with col3:
            st.write(f"Seuil: {criterion.threshold_good:.2f}")
        with col4:
            st.write(f"{criterion.score}/100")

    if decision.deal_breakers:
        st.divider()
        st.error("❌ **Problèmes Bloquants**")
        for db in decision.deal_breakers:
            st.markdown(f"  {db}")

    if decision.warnings:
        st.divider()
        st.warning("⚠️ **Points d'Attention**")
        for warning in decision.warnings:
            st.markdown(f"  {warning}")

    if decision.recommendations:
        st.divider()
        st.info("💡 **Recommandations**")
        for rec in decision.recommendations:
            st.markdown(f"  {rec}")

    st.divider()

    # =========================================================================
    # EXPORT EXCEL
    # =========================================================================
    st.subheader("📥 Export Analyse Complète")

    company_name = financial_data.get("metadata", {}).get("company_name", "Entreprise")
    date_str = datetime.now().strftime("%Y%m%d")

    col1, col2 = st.columns(2)

    with col1:
        if st.button("📊 Générer Export Excel", type="secondary", use_container_width=True):
            with st.spinner("Génération du fichier Excel..."):
                # Convertir decision en dict pour export
                decision_dict = {
                    "decision": {"value": decision.decision.value},
                    "overall_score": decision.overall_score
                }

                # Convertir lbo en dict
                lbo_export = {
                    "acquisition_price": lbo.acquisition_price,
                    "total_debt": lbo.total_debt,
                    "equity_amount": lbo.equity_amount,
                    "debt_layers": [
                        {
                            "name": layer.name,
                            "amount": layer.amount,
                            "interest_rate": layer.interest_rate,
                            "duration_years": layer.duration_years,
                            "grace_period": layer.grace_period
                        }
                        for layer in lbo.debt_layers
                    ]
                }

                norm_export = {
                    "ebitda_bank": norm_data.ebitda_bank,
                    "ebitda_equity": norm_data.ebitda_equity
                }

                excel_file = create_excel_export(
                    stress_results,
                    projections,
                    decision_dict,
                    lbo_export,
                    norm_export,
                    company_name
                )

                st.session_state.excel_export = excel_file
                st.success("✅ Fichier Excel généré!")

    with col2:
        if "excel_export" in st.session_state:
            st.download_button(
                label="💾 Télécharger Excel",
                data=st.session_state.excel_export,
                file_name=f"analyse_lbo_{company_name}_{date_str}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )

    st.divider()

    # Sauvegarder décision
    st.session_state.acquisition_decision = decision

    if st.button("✅ Valider Décision", type="primary", use_container_width=True):
        st.success("✅ Décision validée! Passez à l'onglet 4: Synthèse →")
        st.balloons()
